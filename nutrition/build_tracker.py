#!/usr/bin/env python3
"""Build calorie_tracker.xlsx from meals.csv + reference.csv.

meals.csv is the source of truth. Every total in the workbook is an Excel
formula over the Meals sheet, so editing a meal row in Excel re-computes the
daily, weekly and monthly numbers without re-running this script.

Rows with status "pending" are excluded from every total: nothing counts until
it is confirmed eaten.

Usage:  python3 nutrition/build_tracker.py
"""

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "calorie_tracker.xlsx"

FONT = "Arial"
INPUT_BLUE = Font(name=FONT, size=10, color="0000FF")   # hardcoded input
FORMULA_BLACK = Font(name=FONT, size=10, color="000000")  # formula
TEXT = Font(name=FONT, size=10)
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, size=12, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

HEAD_FILL = PatternFill("solid", fgColor="2F5597")
PENDING_FILL = PatternFill("solid", fgColor="FFFF00")   # needs confirmation
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

KCAL_FMT = "#,##0"
GRAM_FMT = "0.0"
DATE_FMT = "yyyy-mm-dd"
SIGNED_FMT = "+#,##0;-#,##0;0"

# Personal inputs behind every target. These are written into cells on the
# Targets sheet, and BMR / maintenance / calorie target are formulas over those
# cells - so as the weight comes down, update the cell and the whole workbook
# re-derives itself. Change it here only when regenerating from scratch.
PROFILE = {
    "weight_kg": 91.5,
    "height_cm": 162.5,
    "age_years": 33,
    "sex_constant": 5,          # Mifflin-St Jeor: +5 male, -161 female
    "activity_multiplier": 1.45,
    "deficit_kcal": 550,
    "kcal_per_kg_fat": 7700,
    "protein_low_g": 130,
    "protein_high_g": 145,
    "fat_floor_g": 55,
}

CAL_TARGET = "Targets!$B$16"
PROTEIN_LOW = "Targets!$B$20"


def read_csv(name):
    with open(HERE / name, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def style_header(ws, row, ncols, widths=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w


def build_meals(ws, meals):
    ws["A1"] = "Meal log - one row per meal, calorie/protein/fat given as a low-high range"
    ws["A1"].font = TITLE
    ws["A2"] = ('Edit the blue cells only. Add a new meal as a new row directly under the last one, '
                'then re-run nutrition/build_tracker.py so the Daily and Trends sheets pick it up. '
                'Set status to "pending" for anything not yet confirmed eaten - pending rows are '
                'excluded from every total. Yellow rows are awaiting confirmation.')
    ws["A2"].font = NOTE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 42

    headers = ["Date", "Meal", "Items", "Cal low", "Cal high", "Cal mid",
               "Protein low (g)", "Protein high (g)", "Protein mid (g)",
               "Fat low (g)", "Fat high (g)", "Fat mid (g)", "Status", "Notes"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers),
                 widths=[12, 12, 46, 9, 9, 9, 11, 11, 11, 9, 9, 9, 10, 48])

    for i, m in enumerate(meals):
        r = hr + 1 + i
        pending = m["status"].strip().lower() == "pending"
        ws.cell(row=r, column=1, value=dt.date.fromisoformat(m["date"])).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=m["meal"])
        ws.cell(row=r, column=3, value=m["items"])
        ws.cell(row=r, column=4, value=float(m["kcal_low"]))
        ws.cell(row=r, column=5, value=float(m["kcal_high"]))
        ws.cell(row=r, column=6, value=f"=AVERAGE(D{r}:E{r})")
        ws.cell(row=r, column=7, value=float(m["protein_low"]))
        ws.cell(row=r, column=8, value=float(m["protein_high"]))
        ws.cell(row=r, column=9, value=f"=AVERAGE(G{r}:H{r})")
        ws.cell(row=r, column=10, value=float(m["fat_low"]))
        ws.cell(row=r, column=11, value=float(m["fat_high"]))
        ws.cell(row=r, column=12, value=f"=AVERAGE(J{r}:K{r})")
        ws.cell(row=r, column=13, value=m["status"])
        ws.cell(row=r, column=14, value=m["notes"])

        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.font = FORMULA_BLACK if c in (6, 9, 12) else INPUT_BLUE
            if c in (4, 5, 6):
                cell.number_format = KCAL_FMT
            elif c in (7, 8, 9, 10, 11, 12):
                cell.number_format = GRAM_FMT
            elif c in (3, 14):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if pending:
                cell.fill = PENDING_FILL

    ws.freeze_panes = "A5"
    return hr + 1, hr + len(meals)   # first and last data rows


def build_daily(ws, dates, first, last):
    ws["A1"] = "Daily totals - logged meals only (pending rows excluded)"
    ws["A1"].font = TITLE
    ws["A2"] = ("Every cell on this sheet is a formula over the Meals sheet. Do not type over them - "
                "correct the underlying meal row instead. Mid = midpoint of the low-high range, which "
                "is the number to use for trends.")
    ws["A2"].font = NOTE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 28

    headers = ["Date", "Cal low", "Cal high", "Cal mid",
               "Protein low (g)", "Protein high (g)", "Protein mid (g)",
               "Fat low (g)", "Fat high (g)", "Fat mid (g)",
               "Meals logged", "Pending items", "vs target (mid)"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers),
                 widths=[12, 9, 9, 9, 11, 11, 11, 9, 9, 9, 12, 12, 14])

    src = f"Meals!$A${first}:$A${last}"
    stat = f"Meals!$M${first}:$M${last}"

    def sumifs(col):
        return f'=SUMIFS(Meals!${col}${first}:${col}${last},{src},$A{{r}},{stat},"logged")'

    for i, d in enumerate(dates):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=d).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=sumifs("D").format(r=r))
        ws.cell(row=r, column=3, value=sumifs("E").format(r=r))
        ws.cell(row=r, column=4, value=f"=AVERAGE(B{r}:C{r})")
        ws.cell(row=r, column=5, value=sumifs("G").format(r=r))
        ws.cell(row=r, column=6, value=sumifs("H").format(r=r))
        ws.cell(row=r, column=7, value=f"=AVERAGE(E{r}:F{r})")
        ws.cell(row=r, column=8, value=sumifs("J").format(r=r))
        ws.cell(row=r, column=9, value=sumifs("K").format(r=r))
        ws.cell(row=r, column=10, value=f"=AVERAGE(H{r}:I{r})")
        ws.cell(row=r, column=11, value=f'=COUNTIFS({src},$A{r},{stat},"logged")')
        ws.cell(row=r, column=12, value=f'=COUNTIFS({src},$A{r},{stat},"pending")')
        ws.cell(row=r, column=13, value=f"=D{r}-{CAL_TARGET}")

        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.font = FORMULA_BLACK
            if c in (2, 3, 4):
                cell.number_format = KCAL_FMT
            elif c in (5, 6, 7, 8, 9, 10):
                cell.number_format = GRAM_FMT
            elif c == 13:
                cell.number_format = SIGNED_FMT
            else:
                cell.number_format = "0"
            if i % 2:
                cell.fill = BAND_FILL

    ws.freeze_panes = "A5"
    return hr + 1, hr + len(dates)


def build_trends(ws, dates, dfirst, dlast):
    dsrc = f"Daily!$A${dfirst}:$A${dlast}"

    ws["A1"] = "Trends - averages of the daily midpoints"
    ws["A1"].font = TITLE
    ws["A2"] = ("Targets come from the Targets sheet - change them there, not here.")
    ws["A2"].font = NOTE

    ws["A4"] = "Overall"
    ws["A4"].font = Font(name=FONT, size=10, bold=True)
    overall = [
        ("Days logged", f"=COUNT({dsrc})", "0"),
        ("Avg calories/day (mid)", f"=IFERROR(AVERAGE(Daily!$D${dfirst}:$D${dlast}),0)", KCAL_FMT),
        ("Avg protein/day (mid, g)", f"=IFERROR(AVERAGE(Daily!$G${dfirst}:$G${dlast}),0)", GRAM_FMT),
        ("Avg fat/day (mid, g)", f"=IFERROR(AVERAGE(Daily!$J${dfirst}:$J${dlast}),0)", GRAM_FMT),
        ("Avg vs calorie target", f"=IFERROR(B6-{CAL_TARGET},0)", SIGNED_FMT),
        ("Avg deficit vs maintenance", f"=IFERROR(Targets!$B$15-B6,0)", SIGNED_FMT),
        ("Implied loss (kg/week)",
         f"=IFERROR((Targets!$B$15-B6)*7/Targets!$B$11,0)", "0.00"),
        ("Days hitting protein target",
         f'=COUNTIFS(Daily!$G${dfirst}:$G${dlast},">="&{PROTEIN_LOW})', "0"),
        ("Total pending (unconfirmed) items",
         f"=SUM(Daily!$L${dfirst}:$L${dlast})", "0"),
    ]
    for i, (label, formula, fmt) in enumerate(overall):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = TEXT
        c = ws.cell(row=r, column=2, value=formula)
        c.font = FORMULA_BLACK
        c.number_format = fmt
        c.border = BOX
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14

    # ---- Weekly ----
    weeks = sorted({d - dt.timedelta(days=d.weekday()) for d in dates})
    wr = 16
    ws.cell(row=wr - 1, column=1, value="Weekly (week starting Monday)").font = Font(
        name=FONT, size=10, bold=True)
    headers = ["Week start", "Week end", "Days logged", "Avg cal/day (mid)",
               "Avg protein/day (mid, g)", "Avg fat/day (mid, g)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=wr, column=c, value=h)
    style_header(ws, wr, len(headers), widths=[13, 13, 12, 16, 18, 16])

    for i, ws_start in enumerate(weeks):
        r = wr + 1 + i
        ws.cell(row=r, column=1, value=ws_start).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=ws_start + dt.timedelta(days=6)).number_format = DATE_FMT
        _range_block(ws, r, dsrc, dfirst, dlast)

    # ---- Monthly ----
    months = sorted({d.replace(day=1) for d in dates})
    mr = wr + len(weeks) + 3
    ws.cell(row=mr - 1, column=1, value="Monthly").font = Font(name=FONT, size=10, bold=True)
    headers_m = ["Month start", "Month end", "Days logged", "Avg cal/day (mid)",
                 "Avg protein/day (mid, g)", "Avg fat/day (mid, g)"]
    for c, h in enumerate(headers_m, start=1):
        ws.cell(row=mr, column=c, value=h)
    style_header(ws, mr, len(headers_m))

    for i, m_start in enumerate(months):
        r = mr + 1 + i
        nxt = (m_start.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
        ws.cell(row=r, column=1, value=m_start).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=nxt - dt.timedelta(days=1)).number_format = DATE_FMT
        _range_block(ws, r, dsrc, dfirst, dlast)

    last = mr + len(months) + 2
    ws.cell(row=last, column=1,
            value=("Averages cover only the days present in the log; a part-logged day "
                   "(today's dinner still to come) drags its average down."))
    ws.cell(row=last, column=1).font = NOTE


def _range_block(ws, r, dsrc, dfirst, dlast):
    """Days-logged + three averages for the date window in columns A:B of row r."""
    crit = f'{dsrc},">="&$A{r},{dsrc},"<="&$B{r}'
    cells = [
        (3, f"=COUNTIFS({crit})", "0"),
        (4, f"=IFERROR(AVERAGEIFS(Daily!$D${dfirst}:$D${dlast},{crit}),0)", KCAL_FMT),
        (5, f"=IFERROR(AVERAGEIFS(Daily!$G${dfirst}:$G${dlast},{crit}),0)", GRAM_FMT),
        (6, f"=IFERROR(AVERAGEIFS(Daily!$J${dfirst}:$J${dlast},{crit}),0)", GRAM_FMT),
    ]
    for c, formula, fmt in cells:
        cell = ws.cell(row=r, column=c, value=formula)
        cell.font = FORMULA_BLACK
        cell.number_format = fmt
        cell.border = BOX
    for c in (1, 2):
        cell = ws.cell(row=r, column=c)
        cell.font = INPUT_BLUE
        cell.border = BOX



def build_weights(ws, rows):
    """Weigh-in log with a trailing average, which is the number to read.

    A single morning weight moves several hundred grams on water alone; the
    trailing average is what actually tracks fat."""
    ws["A1"] = "Weigh-ins"
    ws["A1"].font = TITLE
    ws["A2"] = ("Weigh same time, same conditions - first thing, after the toilet, before "
                "food or drink. Add a row per weigh-in; the Targets sheet reads the latest "
                "entry automatically. Read the trailing average, not the raw number.")
    ws["A2"].font = NOTE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 42

    headers = ["Date", "Weight (kg)", "Change", "Since start",
               "Trailing avg (7)", "Note"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers), widths=[12, 12, 10, 12, 16, 44])

    first = hr + 1
    for i, row in enumerate(rows):
        r = first + i
        ws.cell(row=r, column=1, value=dt.date.fromisoformat(row["date"])).number_format = DATE_FMT
        ws.cell(row=r, column=2, value=float(row["weight_kg"])).number_format = "0.00"
        ws.cell(row=r, column=3,
                value="" if i == 0 else f"=B{r}-B{r-1}").number_format = "+0.00;-0.00;0.00"
        ws.cell(row=r, column=4, value=f"=B{r}-$B${first}").number_format = "+0.00;-0.00;0.00"
        win = max(first, r - 6)
        ws.cell(row=r, column=5, value=f"=AVERAGE(B{win}:B{r})").number_format = "0.00"
        ws.cell(row=r, column=6, value=row["note"])
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.font = INPUT_BLUE if c in (1, 2, 6) else FORMULA_BLACK
            if c == 6:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    last = first + len(rows) + 1
    ws.cell(row=last, column=1, value=(
        "A drop in the first week is mostly water and glycogen, not fat. Judge the "
        "trend over 2-3 weeks, and change the activity multiplier on Targets only then."))
    ws.cell(row=last, column=1).font = NOTE
    ws.freeze_panes = "A5"


def build_targets(ws):
    """Inputs and the derived calorie/macro targets. Everything downstream
    references these cells, so updating the weight updates the whole workbook."""
    ws["A1"] = "Targets"
    ws["A1"].font = TITLE
    ws["A2"] = ("Edit the yellow input cells only. BMR uses Mifflin-St Jeor, the equation "
                "recommended for this purpose. Everything below the inputs is a formula, so "
                "dropping the weight here re-derives maintenance, the calorie target and the "
                "carb allowance automatically.")
    ws["A2"].font = NOTE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 42

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 66

    def put(r, label, value, fmt, note="", is_input=False, formula=False):
        ws.cell(row=r, column=1, value=label).font = TEXT
        c = ws.cell(row=r, column=2, value=value)
        c.number_format = fmt
        c.border = BOX
        if is_input:
            c.font = INPUT_BLUE
            c.fill = PENDING_FILL
        else:
            c.font = FORMULA_BLACK
        if note:
            n = ws.cell(row=r, column=3, value=note)
            n.font = NOTE
            n.alignment = Alignment(wrap_text=True, vertical="center")

    ws["A4"] = "Inputs"
    ws["A4"].font = Font(name=FONT, size=10, bold=True)
    put(5, "Weight (kg)", "=INDEX(Weights!$B$5:$B$500,COUNT(Weights!$B$5:$B$500))", "0.00",
        "Latest entry on the Weights sheet - add a weigh-in there and everything below "
        "re-derives. Do not type over it.")
    ws["B5"].font = Font(name=FONT, size=10, color="008000")
    put(6, "Height (cm)", PROFILE["height_cm"], "0.0",
        "Source: stated by user.", is_input=True)
    put(7, "Age (years)", PROFILE["age_years"], "0",
        "Source: stated by user (turns 33 on 2026-09-15).", is_input=True)
    put(8, "Sex constant", PROFILE["sex_constant"], "0",
        "Mifflin-St Jeor term: +5 male, -161 female.", is_input=True)
    put(9, "Activity multiplier", PROFILE["activity_multiplier"], "0.00",
        "1.45 = desk job plus 6 lifting sessions/week. Lifting burns less than "
        "cardio-based charts assume; raise it only if the scale says so.", is_input=True)
    put(10, "Daily deficit (kcal)", PROFILE["deficit_kcal"], "#,##0",
        "550/day = 0.5 kg/week. Assumption, not a measurement.", is_input=True)
    put(11, "kcal per kg of body fat", PROFILE["kcal_per_kg_fat"], "#,##0",
        "Standard 7,700 kcal/kg conversion.", is_input=True)

    ws["A13"] = "Derived"
    ws["A13"].font = Font(name=FONT, size=10, bold=True)
    put(14, "BMR (Mifflin-St Jeor)", "=10*B5+6.25*B6-5*B7+B8", "#,##0",
        "Resting burn: what the body uses doing nothing.")
    put(15, "Maintenance (TDEE)", "=B14*B9", "#,##0",
        "Estimate. The scale over 2-3 weeks is the real measurement.")
    put(16, "CALORIE TARGET", "=B15-B10", "#,##0", "The number to eat.")
    ws["A16"].font = Font(name=FONT, size=10, bold=True)
    ws["B16"].font = Font(name=FONT, size=10, bold=True)
    put(17, "Projected loss (kg/week)", "=B10*7/B11", "0.00")
    put(18, "BMI", "=B5/(B6/100)^2", "0.0",
        "Worth a GP or dietitian check alongside this plan.")

    ws["A20"] = "Macros at the calorie target"
    ws["A20"].font = Font(name=FONT, size=10, bold=True)
    put(21, "Protein target low (g)", PROFILE["protein_low_g"], "0",
        "~2.1 g/kg of estimated lean mass - the level that protects muscle in a "
        "deficit while training 6 days.", is_input=True)
    put(22, "Protein target high (g)", PROFILE["protein_high_g"], "0", is_input=True)
    put(23, "Fat floor (g)", PROFILE["fat_floor_g"], "0",
        "~0.6 g/kg. A floor for hormone health, not a goal to hit.", is_input=True)
    put(24, "Carb allowance (g)", "=(B16-B21*4-B23*9)/4", "0",
        "Whatever is left once protein and the fat floor are paid for. Fuels the PPL work.")


def build_reference(ws, rows):
    ws["A1"] = "Per-item working estimates"
    ws["A1"].font = TITLE
    ws["A2"] = ("Photo-based estimates, kept here so the same item is costed the same way every "
                "time. These are assumptions, not measured values - update one and re-cost the "
                "meals that use it.")
    ws["A2"].font = NOTE
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 28

    headers = ["Item", "Cal low", "Cal high", "Cal mid", "Protein", "Notes"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers), widths=[42, 9, 9, 9, 14, 42])

    for i, row in enumerate(rows):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=row["item"])
        ws.cell(row=r, column=2, value=float(row["kcal_low"]))
        ws.cell(row=r, column=3, value=float(row["kcal_high"]))
        ws.cell(row=r, column=4, value=f"=AVERAGE(B{r}:C{r})")
        ws.cell(row=r, column=5, value=row["protein_note"])
        ws.cell(row=r, column=6, value=row["notes"])
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.font = FORMULA_BLACK if c == 4 else INPUT_BLUE
            if c in (2, 3, 4):
                cell.number_format = KCAL_FMT
            elif c in (1, 6):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"


def main():
    meals = read_csv("meals.csv")
    meals.sort(key=lambda m: (m["date"], m["meal"]))
    reference = read_csv("reference.csv")
    dates = sorted({dt.date.fromisoformat(m["date"]) for m in meals})

    wb = Workbook()
    ws_meals = wb.active
    ws_meals.title = "Meals"
    ws_daily = wb.create_sheet("Daily")
    ws_weights = wb.create_sheet("Weights")
    ws_trends = wb.create_sheet("Trends")
    ws_targets = wb.create_sheet("Targets")
    ws_ref = wb.create_sheet("Reference")

    mfirst, mlast = build_meals(ws_meals, meals)
    dfirst, dlast = build_daily(ws_daily, dates, mfirst, mlast)
    build_weights(ws_weights, read_csv("weights.csv"))
    build_trends(ws_trends, dates, dfirst, dlast)
    build_targets(ws_targets)
    build_reference(ws_ref, reference)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # openpyxl writes formulas with no cached values, so force the spreadsheet
    # app to recalculate on open - otherwise every total reads blank until the
    # user touches a cell.
    wb.calculation.fullCalcOnLoad = True

    wb.save(OUT)
    print(f"wrote {OUT} - {len(meals)} meal rows across {len(dates)} days")


if __name__ == "__main__":
    main()
